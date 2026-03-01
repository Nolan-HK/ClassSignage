import sys
import cv2
import pandas as pd
import pymysql
from PyQt5 import QtGui
from PyQt5.QtCore import Qt, QPoint, QTimer, QDateTime
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import QMainWindow, QApplication, QGraphicsDropShadowEffect, QLineEdit, QTableWidgetItem, QWidget,QMessageBox, QHeaderView
from openpyxl.reader.excel import load_workbook

from Login import Ui_LoginWindow
from api_QThread import WorkerThread
from window_check import Ui_window_check
from window_inform import Ui_window_inform
from window_classInfor import Ui_window_classInfor
from window_opendoor import Ui_window_opendoor
from horizontal import Form

# 公共方法类
class publicFun(QWidget):
    @staticmethod
    def __init__(self):
        super().__init__()
        self.forge_link()

    def forge_link(self):
        self._tracking = False
        self._startPos = QPoint()
        self._endPos = QPoint()
        self.result = ""
        self.count = 0

        # 美化
        self.setWindowFlags(Qt.FramelessWindowHint)  # 去边框
        self.setAttribute(Qt.WA_TranslucentBackground)  # 窗口背景透明
        # 更新时间Timer
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.updateDateTime)
        self.timer.start(1000)
        # 摄像头timer
        self.camera_timer = QTimer(self)
        # 插入数据Timer
        self.insert_timer = QTimer(self)
        self.insert_timer.timeout.connect(self.insert_data)


    def data_link(self):
        try:
            conn = pymysql.connect(
                host="localhost",
                user="root",
                password="123456",
                database="Project2_data",  # todo
            )
            return conn
        except Exception as Err:
            QMessageBox.warning(self, "错误", f"{Err}")

    def insert_data(self):
        window_check.showchecked(self, self.data_link())

    def window_1(self):
        classInform.show()
        check.hide()
        inform.hide()
        opendoor.hide()
    def window_2(self):
        check.show()
        classInform.hide()
        inform.hide()
        opendoor.hide()
    def window_3(self):
        inform.show()
        opendoor.hide()
        classInform.hide()
        check.hide()
    def window_4(self):
        opendoor.show()
        inform.hide()
        classInform.hide()
        check.hide()

    def Exit(self):
        exit()

    def updateDateTime(self):
        self.dateTimeEdit.setDateTime(QDateTime.currentDateTime())

    def mouseMoveEvent(self, e: QtGui.QMouseEvent):
        if self._tracking:
            self._endPos = e.pos() - self._startPos
            self.move(self.pos() + self._endPos)

    def mousePressEvent(self, e: QtGui.QMouseEvent):
        if e.button() == Qt.LeftButton:
            self._startPos = QPoint(e.x(), e.y())
            self._tracking = True

    def mouseReleaseEvent(self, e: QtGui.QMouseEvent):
        if e.button() == Qt.LeftButton:
            self._tracking = False
            self._startPos = None
            self._endPos = None

# 公告类
class window_inform(QMainWindow, Ui_window_inform, publicFun):
    def __init__(self):
        super(window_inform, self).__init__(self)
        self.setupUi(self)
        self.dataShow(self.data_link())

        self.BT_exit.clicked.connect(self.Exit)

        self.BT_close.clicked.connect(self.hide_textBrowser)
        self.BT_close.hide()
        self.textBrowser.hide()
        self.BT_1.clicked.connect(self.window_1)
        self.BT_2.clicked.connect(self.window_2)
        self.BT_3.clicked.connect(self.window_3)
        self.BT_4.clicked.connect(self.window_4)

    def dataShow(self, conn):
        query = "SELECT Title, Text, Time FROM inform"
        self.df = pd.read_sql(query, conn)

        # 设置行数
        self.tableWidget.setRowCount(len(self.df.index))

        # 设置列宽
        self.tableWidget.setColumnWidth(0, 600)
        self.tableWidget.setColumnWidth(1, 220)

        # 填入 Title 和 Time 列的数据
        for row in range(len(self.df.index)):
            # 填充 Title 列
            title_item = QTableWidgetItem(str(self.df.iloc[row, self.df.columns.get_loc("Title")]))
            self.tableWidget.setItem(row, 0, title_item)

            # 填充 Time 列
            time_item = QTableWidgetItem(str(self.df.iloc[row, self.df.columns.get_loc("Time")]))
            self.tableWidget.setItem(row, 1, time_item)

            # 设置背景颜色
            if row % 2 == 0:
                title_item.setBackground(QColor(255, 138, 165, 50))
                time_item.setBackground(QColor(255, 138, 165, 50))
            else:
                title_item.setBackground(QColor(98, 218, 255, 50))
                time_item.setBackground(QColor(98, 218, 255, 50))

        self.tableWidget.itemClicked.connect(self.item_clicked)

    # 点击公告显示内容
    def item_clicked(self, item):
        # 启用打开链接的功能
        self.textBrowser.setOpenExternalLinks(True)
        if item.column() == 0:
            self.textBrowser.show()
            self.BT_close.show()
            row = item.row()
            # 填充Text列
            text_data = self.df.iloc[row, self.df.columns.get_loc("Text")]
            self.textBrowser.setHtml(text_data)

    def hide_textBrowser(self):
        self.textBrowser.hide()
        self.BT_close.hide()

# 签到类
class window_check(QMainWindow, Ui_window_check, publicFun):
    def __init__(self):
        super(window_check, self).__init__(self)
        self.setupUi(self)

        # 未签到人数Timer
        self.show_timer = QTimer(self)
        self.show_timer.timeout.connect(self.showuncheck)
        self.show_timer.start(2000)

        # self.OpenCamera()
        self.initChecktable(self.data_link())
        self.thread = WorkerThread()
        self.thread.periodic_function_done.connect(self.getResult)

        self.BT_exit.clicked.connect(self.Exit)
        self.BT_1.clicked.connect(self.window_1)
        self.BT_2.clicked.connect(self.window_2)
        self.BT_2.clicked.connect(self.RunCamera)
        self.BT_3.clicked.connect(self.window_3)
        self.BT_4.clicked.connect(self.window_4)


    def initChecktable(self, conn):
        query = "DELETE FROM checktable;"
        cursor = conn.cursor()
        cursor.execute(query)
        conn.commit()
        cursor.close()

    def OpenCamera(self):
        self.cap = cv2.VideoCapture(0)
        # 初始化前一帧
        ret, self.pre_frame = self.cap.read()

        self.camera_timer.start(40)
        self.camera_timer.timeout.connect(self.RunCamera)

    def RunCamera(self):
        ret, current_frame = self.cap.read()
        frame = cv2.flip(current_frame, 1)
        imgsize = cv2.resize(frame,(1024,720))
# 帧差法检测画面波动
        # 转换为灰度图
        prev_gray = cv2.cvtColor(self.pre_frame, cv2.COLOR_BGR2GRAY)
        current_gray = cv2.cvtColor(current_frame, cv2.COLOR_BGR2GRAY)
        # 计算帧差
        frame_diff = cv2.absdiff(prev_gray, current_gray)
        # 二值化帧差图像
        _, thresholded_diff = cv2.threshold(frame_diff, 30, 255, cv2.THRESH_BINARY)
        # 计算轮廓
        contours, _ = cv2.findContours(thresholded_diff, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        # 计算轮廓面积
        contour_areas = [cv2.contourArea(contour) for contour in contours]
        # 检测到画面变化时
        if any(area > 400 for area in contour_areas):  # 调整面积阈值以适应实际情况
            self.thread.start()  # todo 开启检测进程
            self.insert_timer.start()   # 开启插入数据进程

        # 显示帧差图像
        # cv2.imshow('Frame Difference', thresholded_diff)
        # 更新前一帧
        self.pre_frame = current_frame.copy()

# 显示摄像头画面在界面上
        cvt_img = cv2.cvtColor(imgsize,cv2.COLOR_BGR2RGB)
        self.Image = QtGui.QImage(cvt_img.data,cvt_img.shape[1],cvt_img.shape[0],QtGui.QImage.Format_RGB888)  # 把读取到的视频数据变成QImage形式
        self.label_camera.setPixmap(QtGui.QPixmap.fromImage(self.Image))
        opendoor.label_camera.setPixmap(QtGui.QPixmap.fromImage(self.Image))
        self.Image.save('resource/cache' + '.png', "PNG", 100)
        cv2.waitKey(1)

    def getResult(self, result):
        if result == None:
            pass
        else:
            print(result)
            self.result = result


    def showchecked(self, conn):

    #  向数据库插入已签到的人
        cur = conn.cursor()
        insert_query = f"""INSERT INTO checktable (Student_Name, Number, Major, Image) SELECT s.Student_Name, s.Number, s.Major, s.Image
                            FROM student s JOIN student st ON s.Student_Name = st.Student_Name WHERE s.Student_Name = '{self.result}';"""
        try:
            cur.execute(insert_query)
            conn.commit()
            cur.close()

    # 显示签到人数
            count_ = 0
            count_query = "SELECT COUNT(*) FROM checktable;"
            self.count = pd.read_sql(count_query,conn).iloc[0,0]
            self.label_checkNum.setText(f"已签到 {self.count} 人")

    # 展示已签到人的信息
            query = f"SELECT * FROM checktable WHERE Student_Name = '{self.result}'"
            df = pd.read_sql(query, conn)

            name = df["Student_Name"][0]
            number = df["Number"][0]
            major = df["Major"][0]
            image = df["Image"][0]

            if count_ == self.count:
                pass
            else:
                form = Form(str(image), str(name), str(number), str(major))
                self.verticalLayout_2.addWidget(form)
                count_ += 1
        except:
            pass

    def showuncheck(self):
        index = 0
        self.comboBox.clear()
        self.comboBox.addItem(str(13 - int(self.count)))
        conn = self.data_link()
    # 显示未签到人数
        query = "SELECT Student_Name FROM student WHERE Student_Name NOT IN (SELECT Student_Name FROM checktable);"
        result = pd.read_sql(query, conn)["Student_Name"]
        # 禁用点击选项
        if index != 0:
            self.comboBox.setCurrentIndex(self.selected_index)
        else:
            self.selected_index = index

        self.comboBox.addItems(result.tolist())


# 课程信息类
class window_classInform(QMainWindow, Ui_window_classInfor, publicFun):
    def __init__(self):
        super(window_classInform, self).__init__(self)
        self.setupUi(self)
        self.showClassTable()
        # 间隔timer时间调用一次
        self.timer.timeout.connect(self.showClassInformation)


        self.BT_exit.clicked.connect(self.Exit)
        self.BT_1.clicked.connect(self.window_1)
        self.BT_2.clicked.connect(self.window_2)
        self.BT_3.clicked.connect(self.window_3)
        self.BT_4.clicked.connect(self.window_4)

    def showClassTable(self, classtable_path="ClassTable_12.xlsx"):
        # 打开 Excel 文件
        workbook = load_workbook(classtable_path)
        # 获取第一个工作表
        sheet = workbook.active

        # 获取行列数
        rows = sheet.max_row
        cols = sheet.max_column

        self.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 设置表格的行列数
        self.tableWidget.setRowCount(rows)
        self.tableWidget.setColumnCount(cols)

        # 逐行读取数据并添加到 QTableWidget
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                cell_value = sheet.cell(row=row, column=col).value
                item = QTableWidgetItem(str(cell_value))
                item.setTextAlignment(Qt.AlignCenter)  # 居中显示文本
                # if item is not None:  # todo
                #     item.setBackground(QColor(255, 255, 255))  # 设置为白色
                self.tableWidget.setItem(row - 1, col - 1, item)
                item.setBackground(QColor(0, 255, 255))  # 设置为白色
    def showClassInformation(self):
        conn = self.data_link()
        # current_time = QDateTime.currentDateTime().toString("hh:mm:ss")
        # current_date = QDateTime.currentDateTime().date()
        # 获取星期几
        # day_of_week = current_date.dayOfWeek()
        # select_col = 0 if day_of_week in (Qt.Saturday, Qt.Sunday) else day_of_week

        query = "SELECT Course,Teacher,date,start_time,end_time,image FROM classtable WHERE date = CURRENT_DATE AND CURRENT_TIME BETWEEN start_time AND end_time;"
        df = pd.read_sql(query, conn)
        # start_time_mapping = {
        #     pd.Timedelta("08:10:00"): 1,
        #     pd.Timedelta("10:10:00"): 2,
        #     pd.Timedelta("13:30:00"): 3,
        #     pd.Timedelta("15:30:00"): 4,
        #     pd.Timedelta("18:30:00"): 5,
        #     pd.Timedelta("20:10:00"): 6,
        # }
        try:
            course = df["Course"][0]
            teacher = df["Teacher"][0]
            date = df["date"][0]
            start_time = pd.Timestamp(date) + df["start_time"][0]
            end_time = pd.Timestamp(date) + df["end_time"][0]

            date_time = f"{date} {start_time:%H:%M}-{end_time:%H:%M}"
            # 在右侧填入相应信息
            self.label_course.setText(course)
            # 摄像头窗口顶部显示当前课程
            check.label_course.setText(course)
            self.label_teacher.setText(teacher)
            self.label_image.setPixmap(QtGui.QPixmap(f"{df['image'][0]}"))
            self.label_Class.setText("人工智能32131,人工智能32132")
            self.label_ClassTime.setText(date_time)

            self.tableWidget.findItems(df["Course"][0], Qt.MatchExactly)[0].setSelected(True)   # 选择当前课程所在单元格
        except:
            pass

        conn.close()

# 开门类
class window_opendoor(QMainWindow, Ui_window_opendoor, publicFun):
    def __init__(self):
        super(window_opendoor, self).__init__(self)
        self.setupUi(self)
        self.init_UI()


    def init_UI(self):
        self.groupBox.hide()
        self.BT_viewPasswd.hide()
        self.BT_exit.clicked.connect(self.window_3)
        self.BT_opendoor.clicked.connect(self.groupBox.show)

        self.BT_confirm.clicked.connect(self.verify)
        self.BT_confirm.clicked.connect(self.lineEdit.clear)

        self.BT_cancel.clicked.connect(self.groupBox.hide)
        self.BT_cancel.clicked.connect(self.lineEdit.clear)

        self.BT_hidePasswd.clicked.connect(self.changeView)
        self.BT_viewPasswd.clicked.connect(self.changeView)

    def changeView(self):
        if self.lineEdit.echoMode() == QLineEdit.Password:
            self.lineEdit.setEchoMode(QLineEdit.Normal)
            self.BT_viewPasswd.show()
            self.BT_hidePasswd.hide()
        elif self.lineEdit.echoMode() == QLineEdit.Normal:
            self.lineEdit.setEchoMode(QLineEdit.Password)
            self.BT_viewPasswd.hide()
            self.BT_hidePasswd.show()

    def verify(self):
        conn = self.data_link()
        passwd = self.lineEdit.text()
        query = f"SELECT CASE WHEN COUNT(*) > 0 THEN 1  END as result FROM user WHERE Password = '{passwd}';"
        df = pd.read_sql(query, conn)
        if df["result"][0]:
            QMessageBox.information(self,"提示", "---门已打开---", QMessageBox.Close)
            self.groupBox.hide()
        else:
            QMessageBox.warning(self, "错误", "认证失败！", QMessageBox.Ok)


#  登录类
class LoginWindow(QMainWindow, Ui_LoginWindow):
    def __init__(self, parent=None):
        super(LoginWindow, self).__init__(parent)
        self.setupUi(self)
        self.init()

        self.setWindowFlags(Qt.FramelessWindowHint)  # 去边框
        self.setAttribute(Qt.WA_TranslucentBackground)  # 窗口背景透明
        self.label_Main.setGraphicsEffect(QGraphicsDropShadowEffect(blurRadius=25, xOffset=0, yOffset=0))  # 窗口阴影

    def init(self):
        self.BT_viewPasswd.hide()
        self.BT_exit.clicked.connect(self.Exit)
        self.BT_Login.clicked.connect(self.login_system)

        self.BT_hidePasswd.clicked.connect(self.changeView)
        self.BT_viewPasswd.clicked.connect(self.changeView)

    def changeView(self):
        if self.lineEdit_passwd.echoMode() == QLineEdit.Password:
            self.lineEdit_passwd.setEchoMode(QLineEdit.Normal)
            self.BT_viewPasswd.show()
            self.BT_hidePasswd.hide()
        elif self.lineEdit_passwd.echoMode() == QLineEdit.Normal:
            self.lineEdit_passwd.setEchoMode(QLineEdit.Password)
            self.BT_viewPasswd.hide()
            self.BT_hidePasswd.show()
    def window_login(self):
        inform.show()
        login.hide()
        check.OpenCamera()

    def login_system(self):
        conn = publicFun.data_link(self)
        account = self.lineEdit_account.text()
        passwd = self.lineEdit_passwd.text()
        query = f"SELECT CASE WHEN COUNT(*) > 0 THEN 1  END as result FROM user WHERE Account = '{account}' AND Password = '{passwd}';"
        df = pd.read_sql(query, conn)
        if df["result"][0]:
            self.window_login()
        else:
            QMessageBox.warning(self, "错误", "检查密码或账号是否正确！", QMessageBox.Ok)

    def Exit(self):
        exit()

#  移动窗体函数
    def mouseMoveEvent(self, e: QtGui.QMouseEvent):  # 重写移动事件
        if self._tracking:
            self._endPos = e.pos() - self._startPos
            self.move(self.pos() + self._endPos)

    def mousePressEvent(self, e: QtGui.QMouseEvent):
        if e.button() == Qt.LeftButton:
            self._startPos = QPoint(e.x(), e.y())
            self._tracking = True

    def mouseReleaseEvent(self, e: QtGui.QMouseEvent):
        if e.button() == Qt.LeftButton:
            self._tracking = False
            self._startPos = None
            self._endPos = None


if __name__ == "__main__":
    app = QApplication(sys.argv)
    login = LoginWindow()
    login.show()

    inform = window_inform()
    classInform = window_classInform()
    check = window_check()
    opendoor = window_opendoor()

    sys.exit(app.exec_())
